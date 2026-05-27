"""openpyxl-based Excel report builder."""

from __future__ import annotations

import io
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill


def create_workbook(
    headers: list[str],
    rows: list[dict[str, Any]],
    sheet_title: str = "Report",
) -> bytes:
    """Build an Excel workbook from headers + row dicts and return as bytes.

    Args:
        headers: Column header labels (in order).
        rows: List of row dicts — missing keys produce an empty cell.
        sheet_title: Name of the worksheet tab.

    Returns:
        Raw .xlsx bytes suitable for streaming as an HTTP response.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title  # type: ignore[union-attr]

    # Header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    ws.append(headers)  # type: ignore[union-attr]
    for cell in ws[1]:  # type: ignore[index]
        cell.font = header_font
        cell.fill = header_fill

    ws.freeze_panes = "A2"  # type: ignore[union-attr]

    # Data rows
    for row in rows:
        ws.append([row.get(h) for h in headers])  # type: ignore[union-attr]

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
